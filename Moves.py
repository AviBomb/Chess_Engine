#Importing Dependencies for reading and writing into xlsx files
from openpyxl import load_workbook
wb = load_workbook(filename='final_moveset_pt1.xlsx')
ws = wb['Sheet']
import xlsxwriter 
workbook = xlsxwriter.Workbook('Moves.xlsx') 
worksheet = workbook.add_worksheet() 
#wb1 = load_workbook(filename='final_moveset_pt2.xlsx')
#ws1 = wb1['Sheet']

moves_arr = []
final_moves = []

#Stores the stockfish file holding the weights of all the games that have been played
move_strength=[]

#Stores a 1*65 array which will be passed to the neural network the 65th entry in the array is the weight from the stockfish file the rest 64 are from the game board
fully_processed_array = []

#Stores all the moves of a perticular game within it the format of the game moves isthat of the fully processed array
game_board = []

#Variables Used to Calculate the stockfish value of the game
game=0
move=0

#Generates the list which holds the stocfish file or all the moves' weights required by the neural network
def move_values():
	import csv
	with open('stockfish.csv', 'r') as csvFile:
	    reader = csv.reader(csvFile)
	    for row in reader:
	        move_strength.append(row[1].split())
	# print(move_strength)
	csvFile.close()

#Resets the game board by storing one games values into the ecel file for further reference
def reset():
	move=0;
	worksheet.write(game, 0, game_board) 
	# workbook.close() 
	game_board = []
	game=game+1;

#Creates a Game Board of dimension 1*65 which will be passed to the neural network the 65th entry in the array is the weight from the stockfish file the rest 64 are from the game board
#It then Appends a Move sets final processed array to a game_board which will eventually hold all the moves of a game is reset for every new game
def final_array():
	fully_processed_array = []
	for row in game_board:
		for col in game_board[row]:
				fully_processed_array.append(col)
	fully_processed_array.append(movestrength[game][move])
	move=move+1
	game_board.append(fully_processed_array)

#Reseting the board to it's intial starting position for the next games moves to be recorded
def board_reset ():
	all_moves=[]
	game_board = {'a': [1,6,0,0,0,0,-6,-1],'b': [2,6,0,0,0,0,-6,-2],'c': [3,6,0,0,0,0,-6,-3],'d': [4,6,0,0,0,0,-6,-4],'e': [5,6,0,0,0,0,-6,-5],'f': [3,6,0,0,0,0,-6,-3],'g': [2,6,0,0,0,0,-6,-2],'h': [1,6,0,0,0,0,-6,-1]}
	all_moves.append(game_board)
	return(all_moves)

#Modifying the Board based on the moves being made by the different players
def board (games):
	te=0
	all_moves = board_reset ()
	for move in games:
		print (move)
		if (len(move) == 2):
			for j in range(0,2):
				move[j]=move[j].replace('+', '')
				move[j]=move[j].replace('#', '')
				a=list(move[j])
				if(len(move[j])==2 and move[j].islower() == True):
					for element in all_moves:
						for key in element :
							if(key==a[0]):
								for k in range (0,8):
									if(element[key][k]==6 and j==0 and (int(a[1])==k+2 or ((int(a[1])==k+3) and (k==1) and (element[key][int(a[1])-2]==0)))):
										element[key][k]=0
										element[key][(int(a[1])-1)]=6
										final_moves.append(element)
									if(element[key][k]==-6 and j==1 and (int(a[1])==k or ((int(a[1])==k-1) and (k==6) and (element[key][int(a[1])-1]!=-6)))):
										element[key][k]=0
										element[key][(int(a[1])-1)]=-6
										final_moves.append(element)
				if (len(move[j]) == 4 and move[j].islower() == True):
					a=list(move[j])
					for element in all_moves:
						for key in element :
							if(key==a[2]):
								for k in range (0,8):
									if((k+1)==int(a[3]) and j==0 and element[key][k]!=0):
										te=1
										element[key][k]=6
									if((k+1)==int(a[3]) and j==1 and element[key][k]!=0):
										te=1
										element[key][k]=-6
					for element in all_moves:
						for key in element :
							if(key==a[0]):
								for k in range (0,8):
									if(element[key][k]==6 and j==0 and int(a[3]) == k+2 and te==1):
										element[key][k]=0
										final_moves.append(element)
									if(element[key][k]==-6 and j==1 and int(a[3]) == k and te==1):
										element[key][k]=0
										final_moves.append(element)

				#Pawn Promotion without killing any other piece
				a=list(move[j])
				n=len(a)
				if(len(move[j])==4 and move[j].islower() == False):
					if(a[2]=='='):
						for element in all_moves:
							for key in element :
								if(key==a[0]):
									for k in range (0,8):
										if(element[key][k]==6 and j==0 and int(a[1])==k+2 and k==6 and element[key][int(a[1])-1]==0):
											element[key][k]=0
											if (a[3]=='Q'):
												element[key][(int(a[1])-1)]=4
											if (a[3]=='B'):
												element[key][(int(a[1])-1)]=3
											if (a[3]=='N'):
												element[key][(int(a[1])-1)]=2
											if (a[3]=='R'):
												element[key][(int(a[1])-1)]=1
											final_moves.append(element)
										if(element[key][k]==-6 and j==1 and int(a[1])==k and k==1 and element[key][int(a[1])-1]==0):
											element[key][k]=0
											if (a[3]=='Q'):
												element[key][(int(a[1])-1)]=-4
											if (a[3]=='B'):
												element[key][(int(a[1])-1)]=-3
											if (a[3]=='N'):
												element[key][(int(a[1])-1)]=-2
											if (a[3]=='R'):
												element[key][(int(a[1])-1)]=-1
											final_moves.append(element)

				#Pawn Promotion with killing some other piece
				a=list(move[j])
				n=len(a)
				if (len(move[j]) == 6 and move[j].islower() == False):
					if(a[4]=='='):
						for element in all_moves:
							for key in element :
								if(key==a[0]):
									for k in range (0,8):
										if(element[key][k]==6 and j==0 and int(a[3])==k+2 and k==6):
											element[key][k]=0
											final_moves.append(element)
										if(element[key][k]==-6 and j==1  and int(a[3])==k and k==1):
											element[key][k]=0
											final_moves.append(element)
						for element in all_moves:
							for key in element :
								if(key==a[2]):
									for k in range (0,8):
										if((k+1)==a[3] and j==0):
											if (a[5]=='Q'):
												element[key][(int(a[3])-1)]=4
											if (a[5]=='B'):
												element[key][(int(a[3])-1)]=3
											if (a[5]=='N'):
												element[key][(int(a[3])-1)]=2
											if (a[5]=='R'):
												element[key][(int(a[3])-1)]=1
											final_moves.append(element)
										if((k+1)==a[3] and j==1):
											if (a[5]=='Q'):
												element[key][(int(a[3])-1)]=-4
											if (a[5]=='B'):
												element[key][(int(a[3])-1)]=-3
											if (a[5]=='N'):
												element[key][(int(a[3])-1)]=-2
											if (a[5]=='R'):
												element[key][(int(a[3])-1)]=-1
											final_moves.append(element)

				#King Side Castling 
				if (move[j] == "O-O"):
					for element in all_moves:
						for key in element :
							if(j==0):
								element['e'][0]=0
								element['g'][0]=5
								element['h'][0]=0
								element['f'][0]=1
								final_moves.append(element)
							if(j==1):
								element['e'][7]=0
								element['g'][7]=-5
								element['h'][7]=0
								element['f'][7]=-1
								final_moves.append(element)

				#Queen Side Castling 
				if (move[j] == "O-O-O"):
					for element in all_moves:
						for key in element :
							if(j==0):
								element['e'][0]=0
								element['c'][0]=5
								element['a'][0]=0
								element['d'][0]=1
								final_moves.append(element)
							if(j==1):
								element['e'][7]=0
								element['c'][7]=-5
								element['a'][7]=0
								element['d'][7]=-1
								final_moves.append(element)

				a=list(move[j])
				if(a[0]=='K'):
					if (len(move[j]) == 3):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==5 and j==0):
										element[key][k]=0
										temp=a[1]
										element[temp][(int(a[2])-1)]=5
										final_moves.append(element)
									if(element[key][k]==-5 and j==1):
										element[key][k]=0
										temp=a[1]
										element[temp][(int(a[2])-1)]=-5
										final_moves.append(element)
					if (len(move[j]) == 4):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==5 and j==0):
										element[key][k]=0
										temp=a[2]
										element[temp][(int(a[3])-1)]=5
										final_moves.append(element)
									if(element[key][k]==11 and j==1):
										element[key][k]=0
										temp=a[2]
										element[temp][(int(a[3])-1)]=-5
										final_moves.append(element)


				a=list(move[j])
				if(a[0]=='Q'):
					if (len(move[j]) == 3):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==4 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[1],(int(a[2])-1)]==arr_possibilities):
												element[key][k]=0
												element[a[1]][(int(a[2])-1)]=4
									if(element[key][k]==-4 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[1],(int(a[2])-1)]==arr_possibilities):
												element[key][k]=0
												element[a[1]][(int(a[2])-1)]=-4
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] == 'x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==4 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[2],(int(a[3])-1)]==arr_possibilities):
												element[key][k]=0
												element[a[2]][(int(a[3])-1)]=4
									if(element[key][k]==-4 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[2],(int(a[3])-1)]==arr_possibilities):
												element[key][k]=0
												element[a[2]][(int(a[3])-1)]=-4
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] !='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==4 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],(int(a[3])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[2]][int(a[3]-1)]=4
											if(a[1].isnumeric() == True):
												if([a[2],(int(a[3])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[2]][(int(a[3])-1)]=4
									if(element[key][k]==-4 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],(int(a[3])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[2]][(int(a[3])-1)]=-4
											if(a[1].isnumeric() == True):
												if([a[2],(int(a[3])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[2]][(int(a[3])-1)]=-4
					if (len(move[j]) == 5 and a[2] =='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==4 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],(int(a[4])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[3]][(int(a[4])-1)]=4
											if(a[1].isnumeric() == True):
												if([a[3],(int(a[4])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[3]][(int(a[4])-1)]=4
									if(element[key][k]==-4 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],(int(a[4])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[3]][(int(a[4])-1)]=-4
											if(a[1].isnumeric() == True):
												if([a[3],(int(a[4])-1)]==arr_possibilities):
													element[key][k]=0
													element[a[3]][(int(a[4])-1)]=-4

				a=list(move[j])
				if(a[0]=='B'):
					a=list(move[j])
					if (len(move[j]) == 3):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==3 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[1],(int(a[2])-1)]==arr_possibilities):
												element[key][k]=0
												element[a[1]][(int(a[2])-1)]=3
									if(element[key][k]==-3 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i1 in range (0,8):
											arr.append([key,i1])
										for j1 in range (0,8):
											if (j1==0):
												r='a'
											if (j1==1):
												r='b'
											if (j1==2):
												r='c'
											if (j1==3):
												r='d'
											if (j1==4):
												r='e'
											if (j1==5):
												r='f'
											if (j1==6):
												r='g'
											if (j1==7):
												r='h'
											arr.append([r,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[1],int(a[2])]==arr_possibilities):
												element[key][k]=0
												element[a[1]][int(a[2])]=9
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] == 'x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==3 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[2],int(a[3])]==arr_possibilities):
												element[key][k]=0
												element[a[2]][int(a[3])]=3
									if(element[key][k]==9 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if([a[2],int(a[3])]==arr_possibilities):
												element[key][k]=0
												element[a[2]][int(a[3])]=9
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] !='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==3 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[2]==key):
													element[key][k]=0
													element[a[2]][int(a[3])]=3
											if(a[1].isnumeric() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[3]==k):
													element[key][k]=0
													element[a[2]][int(a[3])]=3
									if(element[key][k]==9 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[2]==key):
													element[key][k]=0
													element[a[2]][int(a[3])]=9
											if(a[1].isnumeric() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[3]==k):
													element[key][k]=0
													element[a[2]][int(a[3])]=9
					a=list(move[j])
					if (len(move[j]) == 5 and a[2] =='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==3 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[3]==key):
													element[key][k]=0
													element[a[3]][int(a[4])]=3
											if(a[1].isnumeric() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[4]==k):
													element[key][k]=0
													element[a[3]][int(a[4])]=3
									if(element[key][k]==9 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for l in range (0,8):
											col=temp+l
											row=k+l
											if (col>=8 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for m in range (0,8):
											col=temp-m
											row=k-m
											if (col<=-1 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for n in range (0,8):
											col=temp-n
											row=k+n
											if (col<=-1 or row>=8):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for o in range (0,8):
											col=temp+o
											row=k-o
											if (col>=8 or row<=-1):
												break
											if (col==0):
												r='a'
											if (col==1):
												r='b'
											if (col==2):
												r='c'
											if (col==3):
												r='d'
											if (col==4):
												r='e'
											if (col==5):
												r='f'
											if (col==6):
												r='g'
											if (col==7):
												r='h'
											arr.append([r,row])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[3]==key):
													element[key][k]=0
													element[a[3]][int(a[4])]=9
											if(a[1].isnumeric() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[4]==k):
													element[key][k]=0
													element[a[3]][int(a[4])]=9
				a=list(move[j])
				if(a[0]=='R'):
					if (len(move[j]) == 3):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==1 and j==0):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if([a[1],int(a[2])]==arr_possibilities):
												element[key][k]=0
												element[a[1]][int(a[2])]=1
									if(element[key][k]==7 and j==1):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if([a[1],int(a[2])]==arr_possibilities):
												element[key][k]=0
												element[a[1]][int(a[2])]=7
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] == 'x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==1 and j==0):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if([a[2],int(a[3])]==arr_possibilities):
												element[key][k]=0
												element[a[2]][int(a[3])]=1
									if(element[key][k]==7 and j==1):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if([a[2],int(a[3])]==arr_possibilities):
												element[key][k]=0
												element[a[2]][int(a[3])]=7
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] !='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==1 and j==0):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[2]==key):
													element[key][k]=0
													element[a[2]][int(a[3])]=1
											if(a[1].isnumeric() == True):
												if([a[2],int(a[3])]==arr_possibilities and int(a[3])==k):
													element[key][k]=0
													element[a[2]][int(a[3])]=1
									if(element[key][k]==7 and j==1):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[2]==key):
													element[key][k]=0
													element[a[2]][int(a[3])]=7
											if(a[1].isnumeric() == True):
												if([a[2],int(a[3])]==arr_possibilities and int(a[3])==k):
													element[key][k]=0
													element[a[2]][int(a[3])]=7
					print(move[j])
					print(list(move[j]))
					a=list(move[j])
					if (len(move[j]) == 5 and a[2] =='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==1 and j==0):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[3]==key):
													element[key][k]=0
													element[a[3]][int(a[4])]=1
											if(a[1].isnumeric() == True):
												if([a[3],int(a[4])]==arr_possibilities and int(a[4])==k):
													element[key][k]=0
													element[a[3]][int(a[4])]=1
									if(element[key][k]==7 and j==1):
										arr=[]
										for i in range (0,8):
											arr.append([key,i])
										for j in range (0,8):
											arr.append([j,k])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[3]==key):
													element[key][k]=0
													element[a[3]][int(a[4])]=7
											if(a[1].isnumeric() == True):
												if([a[3],int(a[4])]==arr_possibilities and int(a[4])==k):
													element[key][k]=0
													element[a[3]][int(a[4])]=7
				a=list(move[j])
				if(a[0]=='N'):
					if (len(move[j]) == 3):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==2 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if([a[1],int(a[2])]==arr_possibilities):
												element[key][k]=0
												element[a[1]][int(a[2])]=2
									if(element[key][k]==8 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if([a[1],int(a[2])]==arr_possibilities):
												element[key][k]=0
												element[a[1]][int(a[2])]=8
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] == 'x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==2 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if([a[2],int(a[3])]==arr_possibilities):
												element[key][k]=0
												element[a[2]][int(a[3])]=2
									if(element[key][k]==8 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if([a[2],int(a[3])]==arr_possibilities):
												element[key][k]=0
												element[a[2]][int(a[3])]=8
					a=list(move[j])
					if (len(move[j]) == 4 and a[1] !='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==2 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[2]==key):
													element[key][k]=0
													element[a[2]][int(a[3])]=2
											if(a[1].isnumeric() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[3]==k):
													element[key][k]=0
													element[a[2]][int(a[3])]=2
									if(element[key][k]==8 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[2],int(a[3])]==arr_possibilities and a[2]==key):
													element[key][k]=0
													element[a[2]][int(a[3])]=8
											if(a[1].isnumeric() == True):
												if([a[2],int(a[3])]==arr_possibilities and int(a[3])==k):
													element[key][k]=0
													element[a[2]][int(a[3])]=8
					a=list(move[j])
					if (len(move[j]) == 5 and a[2] =='x'):
						for element in all_moves:
							for key in element :
								for k in range (0,8):
									if(element[key][k]==2 and j==0):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[3]==key):
													element[key][k]=0
													element[a[3]][int(a[4])]=2
											if(a[1].isnumeric() == True):
												if([a[3],int(a[4])]==arr_possibilities and int(a[4])==k):
													element[key][k]=0
													element[a[3]][int(a[4])]=2
									if(element[key][k]==8 and j==1):
										temp=-1
										if (key=='a'):
											temp=0
										if (key=='b'):
											temp=1
										if (key=='c'):
											temp=2
										if (key=='d'):
											temp=3
										if (key=='e'):
											temp=4
										if (key=='f'):
											temp=5
										if (key=='g'):
											temp=6
										if (key=='h'):
											temp=7
										arr=[]
										if(k+2<=7):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k+2])
										if(k-2>=0):
											if(temp-1>=0):
												temp=temp-1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
											if(temp+1<=7):
												temp=temp+1
												if (temp==0):
													r='a'
												if (temp==1):
													r='b'
												if (temp==2):
													r='c'
												if (temp==3):
													r='d'
												if (temp==4):
													r='e'
												if (temp==5):
													r='f'
												if (temp==6):
													r='g'
												if (temp==7):
													r='h'
												arr.append([r,k-2])
										if(temp+2<=7):
											temp=temp+2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										if(temp-2>=0):
											temp=temp-2
											if (temp==0):
												r='a'
											if (temp==1):
												r='b'
											if (temp==2):
												r='c'
											if (temp==3):
												r='d'
											if (temp==4):
												r='e'
											if (temp==5):
												r='f'
											if (temp==6):
												r='g'
											if (temp==7):
												r='h'
											if(k-1>=0):
												arr.append([r,k-1])
											if(k+1<=7):
												arr.append([r,k+1])
										for arr_possibilities in arr:
											if(a[1].isalpha() == True):
												if([a[3],int(a[4])]==arr_possibilities and a[3]==key):
													element[key][k]=0
													element[a[3]][int(a[4])]=8
											if(a[1].isnumeric() == True):
												if([a[3],int(a[4])]==arr_possibilities and int(a[4])==k):
													element[key][k]=0
													element[a[3]][int(a[4])]=8
					print(all_moves)
					print(final_moves)

# Main Program / Driver Porgram that executes the file
move_values()

for row in ws.rows:
	wm=row[0].value
	bm=row[1].value
	re=row[2].value
	if(type(re) == str or bm == "1/2-1/2" or bm == "1-0" or bm == "0-1" or wm == "1/2-1/2" or wm == "1-0" or wm == "0-1"):
		if(type(re) == str):
			arr = [wm,bm,re]
			moves_arr.append(arr)
		elif(type(re) != str and (bm == "1/2-1/2" or bm == "1-0" or bm == "0-1")):
			arr = [wm,bm]
			moves_arr.append(arr)
		elif(type(re) != str and (wm == "1/2-1/2" or wm == "1-0" or wm == "0-1")):
			arr = [wm]
			moves_arr.append(arr)
		board(moves_arr)
		moves_arr = []
		# continue
		break
	elif(type(re) != str):
		arr = [wm,bm]
		moves_arr.append(arr)
		continue

# for row in ws.rows:
# 	wm=row[0].value
# 	bm=row[1].value
# 	re=row[2].value
# 	if(type(re) == str or bm == "1/2-1/2" or bm == "1-0" or bm == "0-1" or wm == "1/2-1/2" or wm == "1-0" or wm == "0-1"):
# 		arr = [wm,bm,re]
# 		moves_arr.append(arr)
# 		#games.append(moves_arr)
# 		board(moves_arr)
# 		moves_arr = []
# 		break
# 	elif(type(re) != str):
# 		arr = [wm,bm]
# 		moves_arr.append(arr)